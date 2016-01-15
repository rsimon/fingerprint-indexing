from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")


'''
The simplest solution with a lower bound would be something like this:

# Your code:
from openpyxl import load_workbook
filename = 'file_path'
wb = load_workbook(filename, use_iterators=True)
ws = wb.get_sheet_by_name('LOG')

# Solution 1:
for row in ws.iter_rows(row_offset=1):
    # code to execute per row...
Here another way to execute what you describe, with the enumerate function:

# Solution 2:
start, stop = 1, 100    # This will allow you to set a lower and upper limit
for index, row in enumerate(ws.iter_rows()):
    if start < index < stop:
        # code to execute per row...
The index variable keeps count of what row you are on, so it can be used in place of
range or xrange. This method is pretty straightforward and works with iterators unlike
range or slicing, and can be used with just the lower bound too, if desired. Cheers!
'''


'''
Use get_sheet_names() method:

Returns the list of the names of worksheets in the workbook.

Names are returned in the worksheets order.
print wb.get_sheet_names()
You can also get worksheet objects from wb.worksheets:

ws = wb.worksheets[0]
'''
